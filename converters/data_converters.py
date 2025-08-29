# converters/data_converters.py - Complete Enhanced Version

import os
import csv
import json
import uuid
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

import pandas as pd
import xmltodict
import xml.etree.ElementTree as ET

# Enhanced imports for new features
try:
    import yaml  # pip install PyYAML
except ImportError:
    yaml = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    Workbook = load_workbook = dataframe_to_rows = None

try:
    import pyarrow.parquet as pq
    import pyarrow as pa
except ImportError:
    pq = pa = None

# ============== Utils ==============

def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)

def _unique_name(prefix: str, ext: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex}{ext}"

def _sanitize_xml_tag(tag: str) -> str:
    """
    Make a safe XML tag:
    - remove invalid characters
    - replace spaces/illegal chars with underscore
    - ensure it does not start with a digit
    """
    import re
    tag = str(tag).strip()
    tag = re.sub(r"[^A-Za-z0-9_\-\.]", "_", tag)
    if not tag or tag[0].isdigit():
        tag = f"n_{tag}"
    return tag

def _csv_detect_delimiter(file_path: str, default=",") -> str:
    try:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            return dialect.delimiter
    except Exception:
        return default

def _json_to_rows(data: Any) -> List[Dict[str, Any]]:
    """
    Convert arbitrary JSON (list/dict/scalars) to list of flat dict rows.
    Prefer using pandas.json_normalize when nested.
    """
    if isinstance(data, list):
        if not data:
            return []
        # Use json_normalize for nested records
        try:
            df = pd.json_normalize(data, sep=".")
            return df.to_dict(orient="records")
        except Exception:
            return [({"value": json.dumps(item, ensure_ascii=False)} if not isinstance(item, dict) else item) for item in data]
    if isinstance(data, dict):
        try:
            df = pd.json_normalize(data, sep=".")
            return df.to_dict(orient="records")
        except Exception:
            return [data]
    # scalar
    return [{"value": data}]

def _first_record_list_from_xmldict(d: Any) -> Optional[List[Dict[str, Any]]]:
    """
    Try to locate a list[dict] within an xmltodict result.
    """
    if isinstance(d, list):
        # Expect list of dicts
        if d and isinstance(d[0], dict):
            return d  # type: ignore
        return None
    if isinstance(d, dict):
        for k, v in d.items():
            found = _first_record_list_from_xmldict(v)
            if isinstance(found, list):
                return found
    return None

def _rows_to_xml(root_tag: str, record_tag: str, rows: List[Dict[str, Any]]) -> ET.Element:
    root = ET.Element(_sanitize_xml_tag(root_tag))
    for row in rows:
        rec = ET.SubElement(root, _sanitize_xml_tag(record_tag))
        for k, v in row.items():
            tag = _sanitize_xml_tag(k)
            if isinstance(v, (list, dict)):
                child = ET.SubElement(rec, tag)
                child.text = json.dumps(v, ensure_ascii=False)
            else:
                child = ET.SubElement(rec, tag)
                child.text = "" if v is None else str(v)
    return root

# ============== Original Functions (JSON ↔ CSV ↔ XML) ==============

def json_to_csv(in_path: str, out_path: str):
    """JSON -> CSV (handles dict/list/scalars; nested via json_normalize)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    with open(in_path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)
    rows = _json_to_rows(data)
    if rows:
        df = pd.DataFrame(rows)
    else:
        df = pd.DataFrame([])
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    return out_path

def csv_to_json(in_path: str, out_path: str, orient: str = "records"):
    """CSV -> JSON (delimiter auto-detect, preserve strings)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    sep = _csv_detect_delimiter(in_path)
    df = pd.read_csv(in_path, dtype=str, sep=sep, encoding="utf-8-sig")
    # Normalize NaN -> None for cleaner JSON
    data = json.loads(df.to_json(orient=orient, force_ascii=False))
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return out_path

def xml_to_json(in_path: str, out_path: str, force_list: Optional[List[str]] = None):
    """XML -> JSON via xmltodict (optionally force certain tags as lists)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    with open(in_path, "r", encoding="utf-8") as f:
        data_dict = xmltodict.parse(f.read(), force_list=force_list)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data_dict, f, ensure_ascii=False, indent=2)
    return out_path

def json_to_xml(in_path: str, out_path: str, root_tag: str = "root", record_tag: str = "item"):
    """JSON -> XML (nested JSON supported; lists -> repeated record_tag)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    with open(in_path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)

    def build_elem(tag: str, value: Any) -> ET.Element:
        tag = _sanitize_xml_tag(tag)
        if isinstance(value, dict):
            elem = ET.Element(tag)
            for k, v in value.items():
                child = build_elem(k, v)
                elem.append(child)
            return elem
        elif isinstance(value, list):
            # For lists, create multiple 'record_tag' children
            elem = ET.Element(tag)
            for item in value:
                child = build_elem(record_tag, item)
                elem.append(child)
            return elem
        else:
            elem = ET.Element(tag)
            elem.text = "" if value is None else str(value)
            return elem

    if isinstance(data, list):
        root = ET.Element(_sanitize_xml_tag(root_tag))
        for item in data:
            root.append(build_elem(record_tag, item))
    elif isinstance(data, dict):
        root = build_elem(root_tag, data)
    else:
        root = ET.Element(_sanitize_xml_tag(root_tag))
        child = ET.SubElement(root, _sanitize_xml_tag(record_tag))
        child.text = "" if data is None else str(data)

    tree = ET.ElementTree(root)
    tree.write(out_path, encoding="utf-8", xml_declaration=True)
    return out_path

def csv_to_xml(in_path: str, out_path: str, root_tag: str = "data", record_tag: str = "record"):
    """CSV -> XML (delimiter auto-detect; all values treated as strings)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    sep = _csv_detect_delimiter(in_path)
    df = pd.read_csv(in_path, dtype=str, sep=sep, encoding="utf-8-sig")
    rows = df.fillna("").to_dict(orient="records")
    root = _rows_to_xml(root_tag, record_tag, rows)
    ET.ElementTree(root).write(out_path, encoding="utf-8", xml_declaration=True)
    return out_path

def xml_to_csv(in_path: str, out_path: str, record_xpath: Optional[str] = None):
    """
    XML -> CSV
    - If pandas.read_xml is available, try it (handles XSLT/xpath if provided).
    - Fallback: xmltodict -> find first list[dict] -> DataFrame.
    """
    _ensure_dir(Path(out_path).parent.as_posix())

    # Try pandas.read_xml when available
    try:
        # Requires lxml installed; record_xpath optional (e.g., '//record' or '//user')
        if hasattr(pd, "read_xml"):
            if record_xpath:
                df = pd.read_xml(in_path, xpath=record_xpath)
            else:
                # Try to infer a table; if fails, fall back
                df = pd.read_xml(in_path)  # may raise
            df = df.fillna("")
            df.to_csv(out_path, index=False, encoding="utf-8-sig")
            return out_path
    except Exception:
        pass

    # Fallback via xmltodict
    with open(in_path, "r", encoding="utf-8") as f:
        data_dict = xmltodict.parse(f.read())

    records = _first_record_list_from_xmldict(data_dict)
    if records and isinstance(records, list):
        # Ensure rows are dicts; coerce others
        coerced = []
        for r in records:
            if isinstance(r, dict):
                coerced.append(r)
            else:
                coerced.append({"value": r})
        df = pd.json_normalize(coerced, sep=".")
    else:
        # No clear list; dump single-row
        df = pd.json_normalize(data_dict, sep=".")

    df = df.fillna("")
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    return out_path

# ============== Phase 1: YAML & Excel Support ==============

def json_to_yaml(in_path: str, out_path: str):
    """JSON → YAML (configuration files ke liye)"""
    if not yaml:
        raise RuntimeError("PyYAML not installed. Run: pip install PyYAML")
    
    _ensure_dir(Path(out_path).parent.as_posix())
    with open(in_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    with open(out_path, "w", encoding="utf-8") as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True, indent=2)
    return out_path

def yaml_to_json(in_path: str, out_path: str):
    """YAML → JSON (API integration ke liye)"""
    if not yaml:
        raise RuntimeError("PyYAML not installed. Run: pip install PyYAML")
    
    _ensure_dir(Path(out_path).parent.as_posix())
    with open(in_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return out_path

def excel_to_json(in_path: str, out_path: str, sheet_name: str = None):
    """Excel → JSON (multiple sheets support)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    
    # Read Excel file
    if sheet_name:
        df = pd.read_excel(in_path, sheet_name=sheet_name, dtype=str)
    else:
        # Read all sheets
        xlsx_file = pd.ExcelFile(in_path)
        if len(xlsx_file.sheet_names) == 1:
            df = pd.read_excel(in_path, dtype=str)
        else:
            # Multiple sheets → nested JSON
            data = {}
            for sheet in xlsx_file.sheet_names:
                sheet_df = pd.read_excel(in_path, sheet_name=sheet, dtype=str)
                data[sheet] = json.loads(sheet_df.fillna("").to_json(orient="records", force_ascii=False))
            
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return out_path
    
    # Single sheet
    data = json.loads(df.fillna("").to_json(orient="records", force_ascii=False))
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return out_path

def json_to_excel(in_path: str, out_path: str, sheet_name: str = "Sheet1"):
    """JSON → Excel (nested JSON को multiple sheets में)"""
    if not Workbook:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")
    
    _ensure_dir(Path(out_path).parent.as_posix())
    
    with open(in_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    if isinstance(data, dict) and all(isinstance(v, list) for v in data.values()):
        # Multiple sheets scenario
        for sheet_key, sheet_data in data.items():
            ws = wb.create_sheet(title=sheet_key[:31])  # Excel sheet name limit
            if sheet_data and isinstance(sheet_data[0], dict):
                df = pd.DataFrame(sheet_data)
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)
    else:
        # Single sheet
        rows = _json_to_rows(data)
        if rows:
            df = pd.DataFrame(rows)
            ws = wb.create_sheet(title=sheet_name)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
        else:
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["No data"])
    
    wb.save(out_path)
    return out_path

def csv_to_excel(in_path: str, out_path: str, sheet_name: str = "Data"):
    """CSV → Excel (formatting और multiple sheets support)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    
    sep = _csv_detect_delimiter(in_path)
    df = pd.read_csv(in_path, sep=sep, encoding="utf-8-sig")
    
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Optional: Add formatting
        worksheet = writer.sheets[sheet_name]
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    return out_path

def excel_to_csv(in_path: str, out_path: str, sheet_name: str = None):
    """Excel → CSV (specific sheet selection)"""
    _ensure_dir(Path(out_path).parent.as_posix())
    
    if sheet_name:
        df = pd.read_excel(in_path, sheet_name=sheet_name)
    else:
        df = pd.read_excel(in_path)  # First sheet by default
    
    df.to_csv(out_path, index=False, encoding="utf-8-sig")
    return out_path

# ============== Phase 2: Advanced Data Formats ==============

def json_to_parquet(in_path: str, out_path: str):
    """JSON → Parquet (data science optimized)"""
    if not pq or not pa:
        raise RuntimeError("pyarrow required: pip install pyarrow")
    
    _ensure_dir(Path(out_path).parent.as_posix())
    
    with open(in_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    rows = _json_to_rows(data)
    df = pd.DataFrame(rows)
    
    # Convert to Parquet
    table = pa.Table.from_pandas(df)
    pq.write_table(table, out_path)
    return out_path

def parquet_to_json(in_path: str, out_path: str):
    """Parquet → JSON"""
    if not pq:
        raise RuntimeError("pyarrow required: pip install pyarrow")
    
    _ensure_dir(Path(out_path).parent.as_posix())
    
    table = pq.read_table(in_path)
    df = table.to_pandas()
    
    data = json.loads(df.to_json(orient="records", force_ascii=False))
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return out_path

def csv_to_sql_insert(in_path: str, out_path: str, table_name: str = "data_table"):
    """CSV → SQL INSERT statements"""
    _ensure_dir(Path(out_path).parent.as_posix())
    
    sep = _csv_detect_delimiter(in_path)
    df = pd.read_csv(in_path, sep=sep, encoding="utf-8-sig")
    
    with open(out_path, "w", encoding="utf-8") as f:
        # Write CREATE TABLE statement
        columns = []
        for col in df.columns:
            safe_col = _sanitize_xml_tag(col)  # Reuse XML sanitizer for SQL
            columns.append(f"{safe_col} TEXT")
        
        create_stmt = f"CREATE TABLE {table_name} (\n  " + ",\n  ".join(columns) + "\n);\n\n"
        f.write(create_stmt)
        
        # Write INSERT statements
        for _, row in df.iterrows():
            values = []
            for val in row.values:
                if pd.isna(val) or val == "":
                    values.append("NULL")
                else:
                    # Escape single quotes
                    escaped = str(val).replace("'", "''")
                    values.append(f"'{escaped}'")
            
            insert_stmt = f"INSERT INTO {table_name} VALUES ({', '.join(values)});\n"
            f.write(insert_stmt)
    
    return out_path

def analyze_data_structure(in_path: str):
    """Analyze data file structure and suggest best conversion format"""
    ext = Path(in_path).suffix.lower()
    analysis = {
        'format': ext,
        'size_bytes': os.path.getsize(in_path),
        'estimated_rows': 0,
        'estimated_columns': 0,
        'has_nested_data': False,
        'recommended_formats': [],
        'warnings': []
    }
    
    try:
        if ext == '.json':
            with open(in_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            rows = _json_to_rows(data)
            analysis['estimated_rows'] = len(rows)
            if rows:
                analysis['estimated_columns'] = len(rows[0].keys())
                # Check for nested structures
                for row in rows[:5]:  # Sample first 5 rows
                    for value in row.values():
                        if isinstance(value, (dict, list)):
                            analysis['has_nested_data'] = True
                            break
            
            analysis['recommended_formats'] = ['csv', 'excel', 'parquet']
            if analysis['has_nested_data']:
                analysis['recommended_formats'].insert(0, 'xml')
        
        elif ext == '.csv':
            df = pd.read_csv(in_path, nrows=0)  # Just headers
            analysis['estimated_columns'] = len(df.columns)
            
            # Count lines for rows
            with open(in_path, 'r', encoding='utf-8') as f:
                analysis['estimated_rows'] = sum(1 for line in f) - 1  # Minus header
            
            analysis['recommended_formats'] = ['json', 'excel', 'xml', 'parquet']
        
        elif ext in ['.xlsx', '.xls']:
            df = pd.read_excel(in_path, nrows=0)
            analysis['estimated_columns'] = len(df.columns)
            
            # Load full file to count rows (for small files)
            if analysis['size_bytes'] < 50 * 1024 * 1024:  # < 50MB
                full_df = pd.read_excel(in_path)
                analysis['estimated_rows'] = len(full_df)
            
            analysis['recommended_formats'] = ['csv', 'json', 'xml']
    
    except Exception as e:
        analysis['warnings'].append(f"Analysis error: {str(e)}")
    
    return analysis

# ============== Enhanced Class Wrapper ==============

class DataConverters:
    """
    Enhanced backward-compatible wrapper with all new features.
    Prefer calling the functional API with explicit out_path.
    """
    def __init__(self, output_dir: Optional[str] = None):
        self._owns_dir = False
        if output_dir:
            self.output_dir = output_dir
            _ensure_dir(self.output_dir)
        else:
            self.output_dir = tempfile.mkdtemp(prefix="data_conv_")
            self._owns_dir = True

    def __del__(self):
        if getattr(self, "_owns_dir", False):
            try:
                # Do not auto-delete by default to keep outputs; toggle if you want
                pass
            except Exception:
                pass

    # Original JSON <-> CSV
    def json_to_csv(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".csv"))
        return json_to_csv(file_path, out)

    def csv_to_json(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".json"))
        return csv_to_json(file_path, out)

    # Original JSON <-> XML
    def json_to_xml(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".xml"))
        return json_to_xml(file_path, out)

    def xml_to_json(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".json"))
        return xml_to_json(file_path, out)

    # Original CSV <-> XML
    def csv_to_xml(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".xml"))
        return csv_to_xml(file_path, out)

    def xml_to_csv(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".csv"))
        return xml_to_csv(file_path, out)

    # NEW: YAML Support
    def json_to_yaml(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".yaml"))
        return json_to_yaml(file_path, out)

    def yaml_to_json(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".json"))
        return yaml_to_json(file_path, out)

    # NEW: Excel Support
    def excel_to_json(self, file_path: str, sheet_name: str = None) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".json"))
        return excel_to_json(file_path, out, sheet_name=sheet_name)

    def json_to_excel(self, file_path: str, sheet_name: str = "Data") -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".xlsx"))
        return json_to_excel(file_path, out, sheet_name=sheet_name)

    def csv_to_excel(self, file_path: str, sheet_name: str = "Data") -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".xlsx"))
        return csv_to_excel(file_path, out, sheet_name=sheet_name)

    def excel_to_csv(self, file_path: str, sheet_name: str = None) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".csv"))
        return excel_to_csv(file_path, out, sheet_name=sheet_name)

    # NEW: Advanced Formats
    def json_to_parquet(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".parquet"))
        return json_to_parquet(file_path, out)

    def parquet_to_json(self, file_path: str) -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".json"))
        return parquet_to_json(file_path, out)

    def csv_to_sql(self, file_path: str, table_name: str = "data_table") -> str:
        out = os.path.join(self.output_dir, _unique_name("converted", ".sql"))
        return csv_to_sql_insert(file_path, out, table_name=table_name)

    # NEW: Analysis
    def analyze_structure(self, file_path: str) -> Dict[str, Any]:
        return analyze_data_structure(file_path)

# ============== Quick Test Function ==============

def test_available_features():
    """Test which features are available based on installed packages"""
    print("🔧 Data Converter Feature Check")
    print("=" * 40)
    
    features = {
        'Core (JSON/CSV/XML)': True,  # Always available with pandas
        'YAML Support': yaml is not None,
        'Excel Support': Workbook is not None,
        'Parquet Support': pq is not None and pa is not None,
    }
    
    for feature, available in features.items():
        status = "✅" if available else "❌"
        print(f"{status} {feature}")
    
    missing = [f for f, available in features.items() if not available]
    if missing:
        print(f"\n💡 To enable missing features:")
        if not features['YAML Support']:
            print("   pip install PyYAML")
        if not features['Excel Support']:
            print("   pip install openpyxl")
        if not features['Parquet Support']:
            print("   pip install pyarrow")
    
    return sum(features.values())

if __name__ == "__main__":
    available_count = test_available_features()
    print(f"\n📊 {available_count}/4 feature sets available")
    
    if available_count >= 3:
        print("🎉 Data converter is ready for production use!")
    else:
        print("⚠️  Install missing packages for full functionality")
